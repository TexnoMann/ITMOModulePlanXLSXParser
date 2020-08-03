from rooms_parser.rooms_parser import *

import pandas as pd
import numpy as np
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook
import json, re, datetime

def noneConvert(input, num=False):
    if input == None or input=="None" or input=="":
        if num:
            return 0
        else:
            return ""
    return input

class LessonsParser:
    def __init__(self, input_filename, time_config):
        self.__workbook = load_workbook(input_filename, data_only = True)
        self.__sheetname = self.__workbook.sheetnames[0]
        self.__sh = self.__workbook[self.__sheetname]
        self.__lessons_list = []
        self.__time_period_list = []
        tconf_file = open(time_config)
        self.__generateListMergedCells()
        self.__parseTimeConfig(tconf_file)

    def parse(self):
        index_row = 0
        for row in self.__sh.rows:
            first_date_in_table = self.__convertMergedCellToData(row[1])
            if isinstance(first_date_in_table ,datetime.datetime):
                self.__parse_table(index_row, row)
            index_row+=1

    def __parseTimeConfig(self, time_config_file):
        self.__time_config_dict = json.load(time_config_file)
        self.__lesson_time = datetime.datetime.strptime(self.__time_config_dict['lesson_time'],'%H:%M').time()
        for i in range(0,len(self.__time_config_dict['start_times'])):
            self.__time_period_list.append(datetime.datetime.strptime(self.__time_config_dict['start_times'][i],'%H:%M').time())
        # print(self.__time_period_list)

    def __parse_table(self, index_row, row):
        index = 0
        dates = [self.__convertMergedCellToData(v) for v in row]
        last_date_index = 1
        # Find dates, where date is None:
        for i in range(1, len(dates)):
            if dates[i]==None:
                dates[i] = dates[last_date_index]
            else:
                last_date_index = i
        # print(dates)
        check_row_exam = list(self.__sh.rows)[index_row-1]
        exam_id_cell = -1
        for c in range(0,len(check_row_exam)):
            if check_row_exam[c].value=="Экзамен":
                exam_id_cell = c

        last_flow_name =""

        for i in range(index_row+1,index_row+8):
            # print(index_row)
            row = list(self.__sh.rows)[i]
            lessons = [self.__convertMergedCellToData(v) for v in row]
            start_finish_times = self.__stringToSFTime(lessons[0])
            for j in range(1, len(lessons)):
                for k in range(0, len(self.__time_period_list)):
                    if start_finish_times[0] == self.__time_period_list[k] and lessons[j] != None:
                        date = dates[j].strftime('%Y-%m-%d')
                        lesson_info ={"date": date  , "pairIndex": k, "flow": lessons[j], "isExam": 0}
                        self.__lessons_list.append(lesson_info)
                        last_flow_name = lessons[j].split('.')[0]
                        # print(lesson_info)

        if(exam_id_cell != -1):
            exam_date = list(self.__sh.rows)[index_row][exam_id_cell].value.strftime('%Y-%m-%d')
            lesson_info ={"date": exam_date  , "pairIndex": 2, "flow": last_flow_name, "isExam": 1}
            self.__lessons_list.append(lesson_info)
            lesson_info ={"date": exam_date  , "pairIndex": 3, "flow": last_flow_name, "isExam": 1}
            self.__lessons_list.append(lesson_info)

    def save(self, output_filename):
        json_str_out = json.dumps(self.__lessons_list, indent = 4, ensure_ascii=False)
        f = open(output_filename, "w")
        print("[COMPLETE] : Write json output in: ", output_filename)
        f.write(json_str_out)
        f.close()


        # print(date_row)
        # for i in range(index_row+1, index_row+7):
        #     pass

    def __stringToSFTime(self, input_string):
        split_string = re.split(';|,|-| ',input_string)
        time1 = split_string[0]
        time2 = split_string[1]
        dt1 = datetime.datetime.strptime(time1,'%H:%M').time()
        dt2 = datetime.datetime.strptime(time2,'%H:%M').time()
        return (dt1, dt2)

    def __convertMergedCellToData(self, cell, num=False, noneConvertFlag=False, checkMerge=True):
        out= cell.value
        if(checkMerge):
            if type(cell).__name__ == 'MergedCell':
                out= self.__getDataFromMergedCells(cell)
        if noneConvertFlag:
            out = noneConvert(out, num)
        return out



    # Generate self.__merged_cells_list -> list of merged cells groups
    def __generateListMergedCells(self):
        merged_cells = self.__sh.merged_cells.ranges
        for mc in merged_cells:
            f_tuple = mc.bounds
            fc_list=[]
            for i in range(0,len(f_tuple),2):
                fc = (f_tuple[i+1], f_tuple[i])
                fc_list.append(fc)
            self.__merged_cells_list.append(fc_list)

    # Get data from merged cell, need generated self.__merged_cells_list
    def __getDataFromMergedCells(self, cell):
        xy = coordinate_from_string(cell.coordinate)
        col = column_index_from_string(xy[0])
        row = xy[1]
        for c in self.__merged_cells_list:
            if (row, col) in c:
                for p in c:
                    if not p is (row, col):
                        return self.__sh.cell(p[0],p[1]).value
