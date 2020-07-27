import re
import json
import numpy as np
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook

# Basic non class methods
def isNone(input):
    return (input=="None" or input==None)

def isFloat(number):
    try:
        inNumberfloat = float(number)
    except ValueError:
        return False
    return True

def noneConvert(input, num=False):
    if input == None or input=="None" or input=="":
        if num:
            return 0
        else:
            return ""
    return input


class TableParserPlan:
    def __init__(self, input_filename, start_row):
        #  Name for json fields
        self.__op_start_row = start_row
        fc = re.split('\(|\)', input_filename)
        print(fc[1])
        self.__fac = fc[1]
        self.__err_log=""
        self.__meta_field_name = "meta"
        self.__disc_field_name = "disciplines"
        self.__out_dict_json={}
        self.__workbook = load_workbook(input_filename, data_only = True)
        self.__sheetname = self.__workbook.sheetnames[0]
        self.__sh = self.__workbook[self.__sheetname]
        self.__merged_cells_list=[]
        self.__generateListMergedCells()

    def parse(self):
        index = 0
        self.__invalid_cells=0
        for row in self.__sh.rows:
            index+=1
            # Skip initial rows
            if index<self.__op_start_row:
                continue

            op_real = self.__convertMergedCellToData(row[0])
            op_name = self.__convertMergedCellToData(row[1])
            op_supvisor = self.__convertMergedCellToData(row[2])
            op_module = self.__convertMergedCellToData(row[3])
            op_discipl = self.__convertMergedCellToData(row[4], checkMerge=False)
            op_ze = self.__convertMergedCellToData(row[5], num = True, noneConvertFlag=True)
            op_kcp = self.__convertMergedCellToData(row[6], num = True, noneConvertFlag=True)
            # Skip a one column(7)
            op_hours_lec = self.__convertMergedCellToData(row[8], num = True, noneConvertFlag=True)
            op_hours_lab = self.__convertMergedCellToData(row[9], num = True, noneConvertFlag=True)
            op_hours_prac = self.__convertMergedCellToData(row[10], num = True, noneConvertFlag=True)

            op_discipl_periods = row[11:18]
            op_teacher_lec = self.__convertMergedCellToData(row[18], noneConvertFlag=True)
            op_teacher_prac =self.__convertMergedCellToData(row[19], noneConvertFlag=True)
            op_comments = self.__convertMergedCellToData(row[20], noneConvertFlag=True)

            # Check None and invalid fields
            # print(op_discipl)
            ok = True
            if isNone(op_real):
                continue
            elif isNone(op_discipl):
                self.__printError("[VALUE ERROR]", index, op_name, "Дисциплины")
                exit(0)
            else:
                if isNone(op_name):
                    ok=False
                    self.__printError("[VALUE WARNING]", index, op_name, "Наименование ОП")
                if isNone(op_real):
                    ok=False
                    self.__printError("[VALUE WARNING]", index, op_real, "Реализатор ОП")
                if isNone(op_supvisor):
                    ok=False
                    self.__printError("[VALUE WARNING]", index, op_supvisor, "Руководитель ОП")
                if isNone(op_module):
                    ok=False
                    self.__printError("[VALUE WARNING]", index, op_module, "Модуль")

                # Check OP In dict
                if op_name not in self.__out_dict_json:
                    self.__out_dict_json[op_name]={}
                    self.__out_dict_json[op_name][self.__meta_field_name]={}
                    self.__out_dict_json[op_name][self.__disc_field_name]=[]


                self.__out_dict_json[op_name][self.__meta_field_name]["realise"] =  op_real
                self.__out_dict_json[op_name][self.__meta_field_name]["supervisor"] = op_supvisor

                # print(op_teacher_lec)

                # Conver a teacher string to list
                teacher_lec_list = re.split(',|;|/', op_teacher_lec);
                teacher_prac_list = re.split(',|;|/', op_teacher_prac);

                # Calculate a sum of ze in period
                sum_calc_ze = 0.0

                #List of marked period cells
                list_period_rasp =[]
                list_colors_cell =[]


                for cell_rasp in op_discipl_periods:
                    #  Get a cell color
                    color_in_hex = cell_rasp.fill.start_color.index
                    list_colors_cell.append(color_in_hex)
                    cell_rasp = noneConvert(cell_rasp.value)
                    if isFloat(cell_rasp):
                        sum_calc_ze+= float(cell_rasp)
                        list_period_rasp.append(cell_rasp)
                    elif cell_rasp!="":
                        list_period_rasp.append(cell_rasp)
                        sum_calc_ze+=1
                    elif color_in_hex != "00000000" and color_in_hex != 0 and color_in_hex != "FFFFFFFF":
                        list_period_rasp.append(1)
                        sum_calc_ze+=1
                    else:
                        list_period_rasp.append(0)
                if "FFFFFFFF" in list_colors_cell:
                    self.__invalid_cells+=1

                # Check valid disciplines plan.
                valid_disp = ((sum_calc_ze!=0.0 and sum_calc_ze <= np.ceil(op_ze/3))
                              or ("Практика" in op_module or "практика" in op_module))
                if not valid_disp:
                    ok=False
                    self.__printError("[VALUE WARNING]", index, np.ceil(op_ze/3), "calculated ze = "+ str(sum_calc_ze) +", colors: "+str(list_colors_cell))
                    if sum_calc_ze!=0.0:
                        human_sum_calc_ze = (int)(sum_calc_ze) if sum_calc_ze/(1.0*((int)(sum_calc_ze))) == 1.0 else sum_calc_ze
                    else:
                        human_sum_calc_ze = 0
                    self.__err_log+=(self.__fac+": "+op_name + ", " + op_supvisor + ", " + op_discipl+ ", Указано з.е: "+str(op_ze) +", Закрашено ячеек: " + str(human_sum_calc_ze) +"\n")

                valid_disp_str = 1 if valid_disp==True else 0

                new_discipl = {"name": op_discipl, "module": op_module, "ze": op_ze,
                                "kcp": op_kcp, "h_lec": op_hours_lec, "h_lab": op_hours_lab,
                                 "h_prac": op_hours_prac, "ze_in_period": list_period_rasp,
                                  "teachers_lec": teacher_lec_list,
                                 "teachers_prac": teacher_prac_list,
                                 "comments": op_comments,
                                 "valid_disc": ok}

                self.__out_dict_json[op_name][self.__disc_field_name].append(new_discipl)

    def save(self, out_filename):
        # Serializing json
        json_str_out = json.dumps(self.__out_dict_json, indent = 4, ensure_ascii=False)
        f = open(out_filename, "w")
        print("[COMPLETE] : Write json output in: ", out_filename)
        f.write(json_str_out)
        f.close()

    def savelog(self, filename):
        print("[LOG] Count invalid discipl: ", self.__invalid_cells)
        f = open(filename, "a")
        print("[COMPLETE] : Write log output file: ", filename)
        f.write(self.__err_log)
        f.close()



    def __convertMergedCellToData(self, cell, num=False, noneConvertFlag=False, checkMerge=True):
        out= cell.value
        if(checkMerge):
            if type(cell).__name__ == 'MergedCell':
                out= self.__getDataFromMergedCells(cell)
        if noneConvertFlag:
            out = noneConvert(out, num)
        return out



    # Generate self.__merged_cells_list -> list of merged cells groups
    def __generateListMergedCells(self):
        merged_cells = self.__sh.merged_cells.ranges
        for mc in merged_cells:
            f_tuple = mc.bounds
            fc_list=[]
            for i in range(0,len(f_tuple),2):
                fc = (f_tuple[i+1], f_tuple[i])
                fc_list.append(fc)
            self.__merged_cells_list.append(fc_list)

    # Get data from merged cell, need generated self.__merged_cells_list
    def __getDataFromMergedCells(self, cell):
        xy = coordinate_from_string(cell.coordinate)
        col = column_index_from_string(xy[0])
        row = xy[1]
        for c in self.__merged_cells_list:
            if (row, col) in c:
                for p in c:
                    if not p is (row, col):
                        return self.__sh.cell(p[0],p[1]).value

    # Convert data in float if can and check if expected float
    def __convertToFloatIfValid(self, number, where, strong=False):
        if number==None:
            inNumberfloat = ""
        else:
            try:
                inNumberfloat = float(number)
            except ValueError:
                if strong:
                    self.printError("[VALUE WARNING]", index, "non int/float", "Float/Int type")
                    exit(0)
                inNumberfloat = number
        return inNumberfloat

    def __printError(self,type, where, value, need):
        print(type, ": ", "Given value \"", str(value), "\" in row ", str(where), " but expected ", need, ".")
